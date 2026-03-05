"""
ОСНОВНОЙ МОДУЛЬ TELEGRAM-БОТА
Версия 6.1 - СТАБИЛЬНЫЙ РЕЛИЗ (С АКТИВНЫМ ПИНГОМ)
"""

import os
import sys
import asyncio
import logging
import re
import threading
import time
import random
import calendar
import signal
import socket
from datetime import datetime, timezone, timedelta, date
from typing import Optional
from contextlib import asynccontextmanager

# Telegram
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, BotCommand
from telegram.ext import (
    Application, 
    CommandHandler, 
    CallbackQueryHandler, 
    ContextTypes,
    MessageHandler,
    filters,
    ConversationHandler
)

# FastAPI
from fastapi import FastAPI
import uvicorn
import requests

# ========== ПРИНУДИТЕЛЬНЫЙ СБРОС ВЕБХУКА ПРИ СТАРТЕ ==========
BOT_TOKEN = os.getenv("BOT_TOKEN", "")
if BOT_TOKEN:
    try:
        requests.post(
            f"https://api.telegram.org/bot{BOT_TOKEN}/deleteWebhook",
            json={"drop_pending_updates": True},
            timeout=5
        )
        print("🔥 Вебхук принудительно сброшен")
    except Exception as e:
        print(f"⚠️ Не удалось сбросить вебхук: {e}")

from config import VERSION, PORT, RENDER_URL, LOCAL_EXCEL_PATH
from yandex_disk import add_expense, add_income, delete_last, download_from_yandex, get_statistics

# Настройка логгирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# Глобальные переменные
bot_app: Optional[Application] = None
startup_time = datetime.now(timezone.utc)
ping_thread = None
shutdown_event = asyncio.Event()

# Состояния для ConversationHandler
(
    WAITING_EXPENSE_AMOUNT, 
    WAITING_INCOME_AMOUNT,
    WAITING_ARCHIVE_EXPENSE_AMOUNT,
    WAITING_ARCHIVE_INCOME_AMOUNT,
    WAITING_COMPARE_PERIOD1_START,
    WAITING_COMPARE_PERIOD1_END,
    WAITING_COMPARE_PERIOD2_START,
    WAITING_COMPARE_PERIOD2_END,
    WAITING_PERIOD_STATS_START,
    WAITING_PERIOD_STATS_END
) = range(10)

# ========== ДАННЫЕ ==========
ALL_CATEGORIES = [
    "🛒 Продукты", "🏠 Коммуналка", "🚗 Транспорт", "💳 Кредиты",
    "🌿 Зелень", "💊 Лекарства и лечение", "🚬 Сигареты и алко",
    "🐱 Кошка", "🧹 Быт расходники", "🎮 Развлечения и хобби",
    "🔨 Дом/ремонт", "👕 Одежда и обувь", "💇 Красота/Уход", "📦 Другое"
]

PRIORITY_CATEGORIES = [
    "🛒 Продукты", "🚗 Транспорт", "🚬 Сигареты и алко",
    "🏠 Коммуналка", "💳 Кредиты", "🎮 Развлечения и хобби"
]

HIDDEN_CATEGORIES = [cat for cat in ALL_CATEGORIES if cat not in PRIORITY_CATEGORIES]

INCOME_SOURCES = ["💼 Зарплата (Жена)", "💼 Зарплата (Муж)", "💻 Подработка (Муж)"]
PAYERS = ["👩 Жена", "👨 Муж"]
PAYMENT_METHODS = ["💵 Наличные", "💳 Карта Муж", "💳 Карта Жена", "📌 Другое"]

MONTHS_RU = {
    1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
    5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
    9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"
}

# ========== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ==========
def get_moscow_time() -> str:
    """Получить текущее время по Москве"""
    moscow_tz = timezone(timedelta(hours=3))
    return datetime.now(moscow_tz).strftime("%H:%M:%S")

def get_current_date() -> str:
    """Получить текущую дату"""
    return datetime.now(timezone(timedelta(hours=3))).strftime("%d.%m.%y")

def get_current_date_obj() -> date:
    """Получить текущую дату как объект date"""
    return datetime.now(timezone(timedelta(hours=3))).date()

def parse_date(date_str: str) -> Optional[date]:
    """Парсит дату из строки ДД.ММ.ГГ"""
    try:
        return datetime.strptime(date_str, "%d.%m.%y").date()
    except:
        return None

def format_date(date_obj: date) -> str:
    """Форматирует дату в ДД.ММ.ГГ"""
    return date_obj.strftime("%d.%m.%y")

# ========== КЛАВИАТУРЫ ==========
def get_main_keyboard():
    """Главное меню (4 кнопки)"""
    keyboard = [
        [InlineKeyboardButton("💰 Расход", callback_data="expense")],
        [InlineKeyboardButton("💵 Доход", callback_data="income")],
        [InlineKeyboardButton("❌ Удалить", callback_data="delete_last")],
        [InlineKeyboardButton("📊 Статистика", callback_data="stats_menu")]
    ]
    return InlineKeyboardMarkup(keyboard)

def get_categories_keyboard(show_archive=True):
    """Клавиатура с категориями расходов"""
    keyboard = []
    
    for i in range(0, len(PRIORITY_CATEGORIES), 2):
        row = []
        row.append(InlineKeyboardButton(text=PRIORITY_CATEGORIES[i], callback_data=f"cat_{PRIORITY_CATEGORIES[i]}"))
        if i + 1 < len(PRIORITY_CATEGORIES):
            row.append(InlineKeyboardButton(text=PRIORITY_CATEGORIES[i + 1], callback_data=f"cat_{PRIORITY_CATEGORIES[i + 1]}"))
        keyboard.append(row)
    
    if HIDDEN_CATEGORIES:
        keyboard.append([InlineKeyboardButton(text="📋 Другие категории...", callback_data="show_hidden_categories")])
    
    if show_archive:
        keyboard.append([InlineKeyboardButton(text="┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄", callback_data="ignore")])
        keyboard.append([InlineKeyboardButton(text="📅 Архивная запись", callback_data="archive_expense")])
    
    keyboard.append([InlineKeyboardButton(text="🔙 Назад", callback_data="back_main")])
    return InlineKeyboardMarkup(keyboard)

def get_income_sources_keyboard(show_archive=True):
    """Клавиатура с источниками дохода"""
    keyboard = []
    
    for source in INCOME_SOURCES:
        keyboard.append([InlineKeyboardButton(text=source, callback_data=f"source_{source}")])
    
    if show_archive:
        keyboard.append([InlineKeyboardButton(text="┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄", callback_data="ignore")])
        keyboard.append([InlineKeyboardButton(text="📅 Архивная запись", callback_data="archive_income")])
    
    keyboard.append([InlineKeyboardButton(text="🔙 Назад", callback_data="back_main")])
    return InlineKeyboardMarkup(keyboard)

def get_hidden_categories_keyboard():
    """Клавиатура со скрытыми категориями"""
    keyboard = []
    for i in range(0, len(HIDDEN_CATEGORIES), 2):
        row = []
        row.append(InlineKeyboardButton(text=HIDDEN_CATEGORIES[i], callback_data=f"cat_{HIDDEN_CATEGORIES[i]}"))
        if i + 1 < len(HIDDEN_CATEGORIES):
            row.append(InlineKeyboardButton(text=HIDDEN_CATEGORIES[i + 1], callback_data=f"cat_{HIDDEN_CATEGORIES[i + 1]}"))
        keyboard.append(row)
    
    keyboard.append([InlineKeyboardButton(text="🔙 Назад к основным", callback_data="back_to_main_categories")])
    return InlineKeyboardMarkup(keyboard)

def get_payers_keyboard():
    """Клавиатура выбора плательщика"""
    keyboard = [[InlineKeyboardButton(text=p, callback_data=f"payer_{p}")] for p in PAYERS]
    keyboard.append([InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_categories")])
    return InlineKeyboardMarkup(keyboard)

def get_payment_methods_keyboard():
    """Клавиатура выбора способа оплаты"""
    keyboard = []
    for i in range(0, len(PAYMENT_METHODS), 2):
        row = []
        row.append(InlineKeyboardButton(text=PAYMENT_METHODS[i], callback_data=f"method_{PAYMENT_METHODS[i]}"))
        if i + 1 < len(PAYMENT_METHODS):
            row.append(InlineKeyboardButton(text=PAYMENT_METHODS[i + 1], callback_data=f"method_{PAYMENT_METHODS[i + 1]}"))
        keyboard.append(row)
    keyboard.append([InlineKeyboardButton(text="🔙 Назад", callback_data="back_to_payers")])
    return InlineKeyboardMarkup(keyboard)

def get_stats_keyboard():
    """Меню статистики"""
    keyboard = [
        [InlineKeyboardButton("📥 Скачать Excel", callback_data="download_excel")],
        [InlineKeyboardButton("📈 За период", callback_data="stats_period")],
        [InlineKeyboardButton("💰 По категориям", callback_data="stats_categories")],
        [InlineKeyboardButton("📊 Баланс", callback_data="stats_balance")],
        [InlineKeyboardButton("📉 Сравнить периоды", callback_data="stats_compare")],
        [InlineKeyboardButton("🔙 Назад", callback_data="back_main")]
    ]
    return InlineKeyboardMarkup(keyboard)

def get_period_type_keyboard():
    """Выбор типа периода для статистики"""
    keyboard = [
        [InlineKeyboardButton("📅 Конкретные даты", callback_data="period_dates")],
        [InlineKeyboardButton("📆 Месяц", callback_data="period_month")],
        [InlineKeyboardButton("📅 Год", callback_data="period_year")],
        [InlineKeyboardButton("🔙 Назад", callback_data="stats_menu")]
    ]
    return InlineKeyboardMarkup(keyboard)

def get_calendar_keyboard(year: int, month: int, callback_prefix: str):
    """Генерирует клавиатуру-календарь"""
    keyboard = []
    
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1
    
    header = [
        InlineKeyboardButton(f"◀️ {MONTHS_RU[prev_month][:3]}", callback_data=f"{callback_prefix}_month_{prev_year}_{prev_month}"),
        InlineKeyboardButton(f"{MONTHS_RU[month]} {year}", callback_data="ignore"),
        InlineKeyboardButton(f"{MONTHS_RU[next_month][:3]} ▶️", callback_data=f"{callback_prefix}_month_{next_year}_{next_month}")
    ]
    keyboard.append(header)
    
    week_days = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    keyboard.append([InlineKeyboardButton(day, callback_data="ignore") for day in week_days])
    
    cal = calendar.monthcalendar(year, month)
    for week in cal:
        row = []
        for day in week:
            if day == 0:
                row.append(InlineKeyboardButton(" ", callback_data="ignore"))
            else:
                date_str = f"{day:02d}.{month:02d}.{str(year)[-2:]}"
                row.append(InlineKeyboardButton(str(day), callback_data=f"{callback_prefix}_date_{date_str}"))
        keyboard.append(row)
    
    today = get_current_date_obj()
    keyboard.append([InlineKeyboardButton("📅 Сегодня", callback_data=f"{callback_prefix}_date_{format_date(today)}")])
    
    if callback_prefix in ["stats_start", "stats_end", "stats_month", "stats_year"]:
        keyboard.append([InlineKeyboardButton("🔙 Назад", callback_data="back_to_period_type")])
    elif callback_prefix in ["compare1_start", "compare1_end", "compare2_start", "compare2_end"]:
        keyboard.append([InlineKeyboardButton("🔙 Назад", callback_data="stats_compare")])
    else:
        keyboard.append([InlineKeyboardButton("🔙 Назад", callback_data="back_main")])
    
    return InlineKeyboardMarkup(keyboard)

def get_year_keyboard(callback_prefix: str):
    """Клавиатура выбора года"""
    current_year = get_current_date_obj().year
    keyboard = []
    
    for year in range(current_year - 2, current_year + 3):
        if year == current_year:
            keyboard.append([InlineKeyboardButton(f"👉 {year} 👈", callback_data=f"{callback_prefix}_{year}")])
        else:
            keyboard.append([InlineKeyboardButton(str(year), callback_data=f"{callback_prefix}_{year}")])
    
    keyboard.append([InlineKeyboardButton("🔙 Назад", callback_data="back_to_period_type")])
    return InlineKeyboardMarkup(keyboard)

def get_delete_keyboard():
    """Клавиатура удаления"""
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(text="🗑 Последний расход", callback_data="delete_expense")],
        [InlineKeyboardButton(text="🗑 Последний доход", callback_data="delete_income")],
        [InlineKeyboardButton(text="🔙 Назад", callback_data="stats_menu")]
    ])

# ========== ОБРАБОТЧИКИ КОМАНД ==========
async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /start"""
    user = update.effective_user
    logger.info(f"🚀 /start от {user.id}")
    
    await update.message.reply_text(
        f"👋 <b>Добро пожаловать, {user.first_name}!</b>\n\n"
        f"👇 <b>Выберите действие:</b>",
        reply_markup=get_main_keyboard(),
        parse_mode="HTML"
    )

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /help"""
    help_text = """
<b>🤖 КАК ПОЛЬЗОВАТЬСЯ:</b>

💰 <b>РАСХОД:</b>
• Нажмите "💰 Расход"
• Выберите категорию
• Выберите кто платил
• Выберите способ оплаты
• Введите сумму

📅 <b>АРХИВНАЯ ЗАПИСЬ:</b>
• При выборе категории нажмите 
  "📅 Архивная запись" внизу
• Выберите дату в календаре
• Введите сумму

📊 <b>СТАТИСТИКА:</b>
• "📈 За период" - любой диапазон
• "📉 Сравнить периоды" - анализ динамики

❌ <b>УДАЛИТЬ:</b>
• Удаление последней записи
    """
    await update.message.reply_text(help_text, parse_mode="HTML")

async def stats_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /stats"""
    await update.message.reply_text(
        "📊 <b>Меню статистики:</b>",
        reply_markup=get_stats_keyboard(),
        parse_mode="HTML"
    )

async def ping_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /ping"""
    await update.message.reply_text(f"🏓 Pong! Время: {get_moscow_time()}")

async def debug_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик команды /debug"""
    user = update.effective_user
    
    try:
        bot_info = await context.bot.get_me()
        bot_status = f"✅ @{bot_info.username}"
    except:
        bot_status = "❌ Ошибка подключения"
    
    response = (
        f"🔧 ДИАГНОСТИКА v{VERSION}\n"
        f"━━━━━━━━━━━━━━━━\n"
        f"🤖 Статус: {bot_status}\n"
        f"🕐 Время: {get_moscow_time()}\n"
        f"📅 Дата: {get_current_date()}\n"
        f"👤 Ваш ID: {user.id}\n"
        f"📊 Режим: Архив + Сравнение\n"
        f"━━━━━━━━━━━━━━━━\n"
        f"💡 Если бот не отвечает:\n"
        f"• Напишите /start\n"
        f"• Подождите 10 секунд\n"
        f"• Попробуйте снова"
    )
    
    await update.message.reply_text(response)

# ========== ОБРАБОТЧИКИ КОЛЛБЭКОВ ==========
async def button_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик нажатий на inline-кнопки"""
    query = update.callback_query
    await query.answer()
    
    data = query.data
    user_id = query.from_user.id
    
    logger.info(f"🔘 Кнопка: {data} от {user_id}")
    
    if data == "ignore":
        return ConversationHandler.END
    
    if data == "back_main":
        await query.edit_message_text(
            "Выберите действие:",
            reply_markup=get_main_keyboard()
        )
        return ConversationHandler.END
    
    elif data == "expense":
        await query.edit_message_text(
            "📌 <b>Выберите категорию расхода:</b>",
            reply_markup=get_categories_keyboard(show_archive=True),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data == "show_hidden_categories":
        await query.edit_message_text(
            "📌 <b>Дополнительные категории:</b>",
            reply_markup=get_hidden_categories_keyboard(),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data == "back_to_main_categories":
        await query.edit_message_text(
            "📌 <b>Выберите категорию расхода:</b>",
            reply_markup=get_categories_keyboard(show_archive=True),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data.startswith("cat_"):
        category = data[4:]
        context.user_data["category"] = category
        context.user_data["is_archive"] = False
        await query.edit_message_text(
            "👤 <b>Кто платил?</b>",
            reply_markup=get_payers_keyboard(),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data == "archive_expense":
        context.user_data["is_archive"] = True
        await query.edit_message_text(
            "📅 <b>Выберите дату расхода:</b>",
            reply_markup=get_calendar_keyboard(
                get_current_date_obj().year,
                get_current_date_obj().month,
                "archive_expense"
            ),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data.startswith("archive_expense_date_"):
        selected_date = data.replace("archive_expense_date_", "")
        context.user_data["archive_date"] = selected_date
        await query.edit_message_text(
            f"📌 <b>Выберите категорию для {selected_date}:</b>",
            reply_markup=get_categories_keyboard(show_archive=False),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data.startswith("archive_expense_month_"):
        parts = data.split("_")
        year = int(parts[3])
        month = int(parts[4])
        await query.edit_message_text(
            "📅 <b>Выберите дату:</b>",
            reply_markup=get_calendar_keyboard(year, month, "archive_expense"),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data == "income":
        await query.edit_message_text(
            "💵 <b>Выберите источник дохода:</b>",
            reply_markup=get_income_sources_keyboard(show_archive=True),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data.startswith("source_"):
        source = data[7:]
        context.user_data["source"] = source
        context.user_data["is_archive"] = False
        await query.edit_message_text(
            "💰 <b>Введите сумму дохода</b>\n(только цифры, например: 50000)",
            parse_mode="HTML"
        )
        return WAITING_INCOME_AMOUNT
    
    elif data == "archive_income":
        context.user_data["is_archive"] = True
        await query.edit_message_text(
            "📅 <b>Выберите дату дохода:</b>",
            reply_markup=get_calendar_keyboard(
                get_current_date_obj().year,
                get_current_date_obj().month,
                "archive_income"
            ),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data.startswith("archive_income_date_"):
        selected_date = data.replace("archive_income_date_", "")
        context.user_data["archive_date"] = selected_date
        await query.edit_message_text(
            f"💵 <b>Выберите источник дохода для {selected_date}:</b>",
            reply_markup=get_income_sources_keyboard(show_archive=False),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data.startswith("archive_income_month_"):
        parts = data.split("_")
        year = int(parts[3])
        month = int(parts[4])
        await query.edit_message_text(
            "📅 <b>Выберите дату:</b>",
            reply_markup=get_calendar_keyboard(year, month, "archive_income"),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data.startswith("payer_"):
        payer = data[6:]
        context.user_data["payer"] = payer
        await query.edit_message_text(
            "💳 <b>Способ оплаты:</b>",
            reply_markup=get_payment_methods_keyboard(),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data == "back_to_categories":
        await query.edit_message_text(
            "📌 <b>Выберите категорию расхода:</b>",
            reply_markup=get_categories_keyboard(show_archive=True),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data == "back_to_payers":
        await query.edit_message_text(
            "👤 <b>Кто платил?</b>",
            reply_markup=get_payers_keyboard(),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data.startswith("method_"):
        method = data[7:]
        context.user_data["method"] = method
        
        if context.user_data.get("is_archive", False):
            archive_date = context.user_data.get("archive_date", get_current_date())
            await query.edit_message_text(
                f"📅 Дата: {archive_date}\n"
                f"💰 <b>Введите сумму расхода</b>\n(только цифры, например: 1500)",
                parse_mode="HTML"
            )
            return WAITING_ARCHIVE_EXPENSE_AMOUNT
        else:
            await query.edit_message_text(
                "💰 <b>Введите сумму расхода</b>\n(только цифры, например: 1500)",
                parse_mode="HTML"
            )
            return WAITING_EXPENSE_AMOUNT
    
    elif data == "stats_menu":
        await query.edit_message_text(
            "📊 <b>Меню статистики:</b>",
            reply_markup=get_stats_keyboard(),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data == "stats_period":
        await query.edit_message_text(
            "📅 <b>Выберите тип периода:</b>",
            reply_markup=get_period_type_keyboard(),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data == "back_to_period_type":
        await query.edit_message_text(
            "📅 <b>Выберите тип периода:</b>",
            reply_markup=get_period_type_keyboard(),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data == "period_dates":
        context.user_data["stats_type"] = "dates"
        await query.edit_message_text(
            "📅 <b>Выберите НАЧАЛЬНУЮ дату:</b>",
            reply_markup=get_calendar_keyboard(
                get_current_date_obj().year,
                get_current_date_obj().month,
                "stats_start"
            ),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data.startswith("stats_start_date_"):
        start_date = data.replace("stats_start_date_", "")
        context.user_data["stats_start"] = start_date
        await query.edit_message_text(
            f"📅 Начальная дата: {start_date}\n\n"
            f"📅 <b>Выберите КОНЕЧНУЮ дату:</b>",
            reply_markup=get_calendar_keyboard(
                get_current_date_obj().year,
                get_current_date_obj().month,
                "stats_end"
            ),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data.startswith("stats_start_month_"):
        parts = data.split("_")
        year = int(parts[3])
        month = int(parts[4])
        await query.edit_message_text(
            "📅 <b>Выберите НАЧАЛЬНУЮ дату:</b>",
            reply_markup=get_calendar_keyboard(year, month, "stats_start"),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data.startswith("stats_end_date_"):
        end_date = data.replace("stats_end_date_", "")
        start_date = context.user_data.get("stats_start")
        
        if start_date:
            await query.edit_message_text("⏳ Считаю статистику...")
            stats_text = get_statistics_period(start_date, end_date)
            await query.edit_message_text(
                f"📊 <b>Статистика за период:</b>\n"
                f"📅 {start_date} - {end_date}\n\n{stats_text}",
                reply_markup=get_stats_keyboard(),
                parse_mode="HTML"
            )
            context.user_data.clear()
        return ConversationHandler.END
    
    elif data.startswith("stats_end_month_"):
        parts = data.split("_")
        year = int(parts[3])
        month = int(parts[4])
        await query.edit_message_text(
            "📅 <b>Выберите КОНЕЧНУЮ дату:</b>",
            reply_markup=get_calendar_keyboard(year, month, "stats_end"),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data == "period_month":
        context.user_data["stats_type"] = "month"
        await query.edit_message_text(
            "📆 <b>Выберите месяц:</b>",
            reply_markup=get_calendar_keyboard(
                get_current_date_obj().year,
                get_current_date_obj().month,
                "stats_month"
            ),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data.startswith("stats_month_date_"):
        selected_date = data.replace("stats_month_date_", "")
        date_obj = parse_date(selected_date)
        if date_obj:
            start_date = date_obj.replace(day=1)
            last_day = calendar.monthrange(date_obj.year, date_obj.month)[1]
            end_date = date_obj.replace(day=last_day)
            
            await query.edit_message_text("⏳ Считаю статистику...")
            stats_text = get_statistics_period(
                format_date(start_date),
                format_date(end_date)
            )
            await query.edit_message_text(
                f"📊 <b>Статистика за {MONTHS_RU[date_obj.month]} {date_obj.year}:</b>\n\n{stats_text}",
                reply_markup=get_stats_keyboard(),
                parse_mode="HTML"
            )
            context.user_data.clear()
        return ConversationHandler.END
    
    elif data.startswith("stats_month_month_"):
        parts = data.split("_")
        year = int(parts[3])
        month = int(parts[4])
        await query.edit_message_text(
            "📆 <b>Выберите месяц:</b>",
            reply_markup=get_calendar_keyboard(year, month, "stats_month"),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data == "period_year":
        context.user_data["stats_type"] = "year"
        await query.edit_message_text(
            "📅 <b>Выберите год:</b>",
            reply_markup=get_year_keyboard("stats_year"),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data.startswith("stats_year_"):
        year = int(data.replace("stats_year_", ""))
        start_date = date(year, 1, 1)
        end_date = date(year, 12, 31)
        
        await query.edit_message_text("⏳ Считаю статистику...")
        stats_text = get_statistics_period(
            format_date(start_date),
            format_date(end_date)
        )
        await query.edit_message_text(
            f"📊 <b>Статистика за {year} год:</b>\n\n{stats_text}",
            reply_markup=get_stats_keyboard(),
            parse_mode="HTML"
        )
        context.user_data.clear()
        return ConversationHandler.END
    
    elif data == "stats_compare":
        context.user_data["compare_step"] = "period1_start"
        await query.edit_message_text(
            "📊 <b>Сравнение периодов</b>\n\n"
            "📅 Выберите НАЧАЛО ПЕРВОГО периода:",
            reply_markup=get_calendar_keyboard(
                get_current_date_obj().year,
                get_current_date_obj().month,
                "compare1_start"
            ),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data.startswith("compare1_start_date_"):
        start1 = data.replace("compare1_start_date_", "")
        context.user_data["compare_start1"] = start1
        context.user_data["compare_step"] = "period1_end"
        await query.edit_message_text(
            f"📅 Первый период: начало {start1}\n\n"
            f"📅 Выберите КОНЕЦ ПЕРВОГО периода:",
            reply_markup=get_calendar_keyboard(
                get_current_date_obj().year,
                get_current_date_obj().month,
                "compare1_end"
            ),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data.startswith("compare1_start_month_"):
        parts = data.split("_")
        year = int(parts[3])
        month = int(parts[4])
        await query.edit_message_text(
            "📅 Выберите НАЧАЛО ПЕРВОГО периода:",
            reply_markup=get_calendar_keyboard(year, month, "compare1_start"),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data.startswith("compare1_end_date_"):
        end1 = data.replace("compare1_end_date_", "")
        context.user_data["compare_end1"] = end1
        context.user_data["compare_step"] = "period2_start"
        await query.edit_message_text(
            f"📅 Первый период: {context.user_data['compare_start1']} - {end1}\n\n"
            f"📅 Выберите НАЧАЛО ВТОРОГО периода:",
            reply_markup=get_calendar_keyboard(
                get_current_date_obj().year,
                get_current_date_obj().month,
                "compare2_start"
            ),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data.startswith("compare1_end_month_"):
        parts = data.split("_")
        year = int(parts[3])
        month = int(parts[4])
        await query.edit_message_text(
            "📅 Выберите КОНЕЦ ПЕРВОГО периода:",
            reply_markup=get_calendar_keyboard(year, month, "compare1_end"),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data.startswith("compare2_start_date_"):
        start2 = data.replace("compare2_start_date_", "")
        context.user_data["compare_start2"] = start2
        context.user_data["compare_step"] = "period2_end"
        await query.edit_message_text(
            f"📅 Второй период: начало {start2}\n\n"
            f"📅 Выберите КОНЕЦ ВТОРОГО периода:",
            reply_markup=get_calendar_keyboard(
                get_current_date_obj().year,
                get_current_date_obj().month,
                "compare2_end"
            ),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data.startswith("compare2_start_month_"):
        parts = data.split("_")
        year = int(parts[3])
        month = int(parts[4])
        await query.edit_message_text(
            "📅 Выберите НАЧАЛО ВТОРОГО периода:",
            reply_markup=get_calendar_keyboard(year, month, "compare2_start"),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data.startswith("compare2_end_date_"):
        end2 = data.replace("compare2_end_date_", "")
        
        start1 = context.user_data.get("compare_start1")
        end1 = context.user_data.get("compare_end1")
        start2 = context.user_data.get("compare_start2")
        
        if not all([start1, end1, start2, end2]):
            await query.edit_message_text(
                "❌ Ошибка: не выбраны все даты",
                reply_markup=get_stats_keyboard()
            )
            context.user_data.clear()
            return ConversationHandler.END
        
        await query.edit_message_text("⏳ Сравниваю периоды...")
        compare_text = compare_periods(start1, end1, start2, end2)
        await query.edit_message_text(
            f"📊 <b>Сравнение периодов</b>\n\n"
            f"Период 1: {start1} - {end1}\n"
            f"Период 2: {start2} - {end2}\n\n{compare_text}",
            reply_markup=get_stats_keyboard(),
            parse_mode="HTML"
        )
        
        context.user_data.clear()
        return ConversationHandler.END
    
    elif data.startswith("compare2_end_month_"):
        parts = data.split("_")
        year = int(parts[3])
        month = int(parts[4])
        await query.edit_message_text(
            "📅 Выберите КОНЕЦ ВТОРОГО периода:",
            reply_markup=get_calendar_keyboard(year, month, "compare2_end"),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data == "download_excel":
        await query.edit_message_text("⏬ Скачиваю файл...")
        
        if download_from_yandex():
            try:
                with open(LOCAL_EXCEL_PATH, "rb") as f:
                    await context.bot.send_document(
                        chat_id=query.message.chat_id,
                        document=f,
                        filename="budget.xlsx",
                        caption="📁 Ваш файл budget.xlsx"
                    )
                await query.message.reply_text(
                    "Выберите действие:",
                    reply_markup=get_stats_keyboard()
                )
            except Exception as e:
                logger.error(f"Ошибка отправки файла: {e}")
                await query.message.reply_text(
                    "❌ Ошибка при отправке",
                    reply_markup=get_stats_keyboard()
                )
        else:
            await query.message.reply_text(
                "❌ Не удалось скачать файл",
                reply_markup=get_stats_keyboard()
            )
        return ConversationHandler.END
    
    elif data == "stats_categories":
        await query.edit_message_text("⏳ Считаю...")
        stats_text = get_statistics(by_categories=True)
        await query.edit_message_text(
            f"📊 <b>Расходы по категориям:</b>\n\n{stats_text}",
            reply_markup=get_stats_keyboard(),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data == "stats_balance":
        await query.edit_message_text("⏳ Считаю...")
        balance_text = get_statistics(balance=True)
        await query.edit_message_text(
            f"💰 <b>Текущий баланс:</b>\n\n{balance_text}",
            reply_markup=get_stats_keyboard(),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data == "delete_last":
        await query.edit_message_text(
            "❓ <b>Что удалить?</b>",
            reply_markup=get_delete_keyboard(),
            parse_mode="HTML"
        )
        return ConversationHandler.END
    
    elif data == "delete_expense":
        result = delete_last("Расходы")
        await query.message.reply_text(result)
        await query.message.reply_text(
            "Выберите действие:",
            reply_markup=get_stats_keyboard()
        )
        return ConversationHandler.END
    
    elif data == "delete_income":
        result = delete_last("Доходы")
        await query.message.reply_text(result)
        await query.message.reply_text(
            "Выберите действие:",
            reply_markup=get_stats_keyboard()
        )
        return ConversationHandler.END
    
    return ConversationHandler.END

# ========== ОБРАБОТЧИКИ СООБЩЕНИЙ ==========
async def handle_expense_amount(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик ввода суммы расхода (текущая дата)"""
    text = update.message.text.strip()
    amount_str = re.sub(r"[^\d.,]", "", text).replace(",", ".")
    
    try:
        amount = float(amount_str)
        if amount <= 0 or amount > 1_000_000:
            raise ValueError
    except ValueError:
        await update.message.reply_text("❌ Введите корректное число (от 1 до 1 000 000):")
        return WAITING_EXPENSE_AMOUNT
    
    category = context.user_data.get("category")
    payer = context.user_data.get("payer")
    method = context.user_data.get("method")
    
    if not all([category, payer, method]):
        await update.message.reply_text(
            "❌ Ошибка сессии. Начните заново.",
            reply_markup=get_main_keyboard()
        )
        context.user_data.clear()
        return ConversationHandler.END
    
    result = add_expense(category, amount, payer, method)
    await update.message.reply_text(result)
    await update.message.reply_text(
        "👇 Выберите следующее действие:",
        reply_markup=get_main_keyboard()
    )
    
    context.user_data.clear()
    return ConversationHandler.END

async def handle_archive_expense_amount(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик ввода суммы расхода (архивная дата)"""
    text = update.message.text.strip()
    amount_str = re.sub(r"[^\d.,]", "", text).replace(",", ".")
    
    try:
        amount = float(amount_str)
        if amount <= 0 or amount > 1_000_000:
            raise ValueError
    except ValueError:
        await update.message.reply_text("❌ Введите корректное число (от 1 до 1 000 000):")
        return WAITING_ARCHIVE_EXPENSE_AMOUNT
    
    category = context.user_data.get("category")
    payer = context.user_data.get("payer")
    method = context.user_data.get("method")
    archive_date = context.user_data.get("archive_date", get_current_date())
    
    if not all([category, payer, method]):
        await update.message.reply_text(
            "❌ Ошибка сессии. Начните заново.",
            reply_markup=get_main_keyboard()
        )
        context.user_data.clear()
        return ConversationHandler.END
    
    result = add_expense(category, amount, payer, method)
    await update.message.reply_text(f"✅ {result} (дата: {archive_date})")
    await update.message.reply_text(
        "👇 Выберите следующее действие:",
        reply_markup=get_main_keyboard()
    )
    
    context.user_data.clear()
    return ConversationHandler.END

async def handle_income_amount(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик ввода суммы дохода (текущая дата)"""
    text = update.message.text.strip()
    amount_str = re.sub(r"[^\d.,]", "", text).replace(",", ".")
    
    try:
        amount = float(amount_str)
        if amount <= 0 or amount > 10_000_000:
            raise ValueError
    except ValueError:
        await update.message.reply_text("❌ Введите корректное число (от 1 до 10 000 000):")
        return WAITING_INCOME_AMOUNT
    
    source = context.user_data.get("source")
    
    if not source:
        await update.message.reply_text(
            "❌ Ошибка сессии. Начните заново.",
            reply_markup=get_main_keyboard()
        )
        context.user_data.clear()
        return ConversationHandler.END
    
    payer = "Муж" if "Муж" in source else "Жена"
    result = add_income(source, amount, payer)
    
    await update.message.reply_text(result)
    await update.message.reply_text(
        "👇 Выберите следующее действие:",
        reply_markup=get_main_keyboard()
    )
    
    context.user_data.clear()
    return ConversationHandler.END

async def handle_archive_income_amount(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик ввода суммы дохода (архивная дата)"""
    text = update.message.text.strip()
    amount_str = re.sub(r"[^\d.,]", "", text).replace(",", ".")
    
    try:
        amount = float(amount_str)
        if amount <= 0 or amount > 10_000_000:
            raise ValueError
    except ValueError:
        await update.message.reply_text("❌ Введите корректное число (от 1 до 10 000 000):")
        return WAITING_ARCHIVE_INCOME_AMOUNT
    
    source = context.user_data.get("source")
    archive_date = context.user_data.get("archive_date", get_current_date())
    
    if not source:
        await update.message.reply_text(
            "❌ Ошибка сессии. Начните заново.",
            reply_markup=get_main_keyboard()
        )
        context.user_data.clear()
        return ConversationHandler.END
    
    payer = "Муж" if "Муж" in source else "Жена"
    result = add_income(source, amount, payer)
    
    await update.message.reply_text(f"✅ {result} (дата: {archive_date})")
    await update.message.reply_text(
        "👇 Выберите следующее действие:",
        reply_markup=get_main_keyboard()
    )
    
    context.user_data.clear()
    return ConversationHandler.END

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отмена действия"""
    await update.message.reply_text(
        "Действие отменено",
        reply_markup=get_main_keyboard()
    )
    context.user_data.clear()
    return ConversationHandler.END

async def handle_unknown(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработчик неизвестных сообщений"""
    await update.message.reply_text(
        "❓ Используйте кнопки меню 👇",
        reply_markup=get_main_keyboard()
    )

# ========== ФУНКЦИИ ДЛЯ СТАТИСТИКИ ==========
def get_statistics_period(start_date: str, end_date: str) -> str:
    """Получает статистику за период"""
    try:
        if not download_from_yandex():
            return "❌ Не удалось скачать файл"
        
        from openpyxl import load_workbook
        
        wb = load_workbook(LOCAL_EXCEL_PATH, data_only=True)
        
        income_total = 0
        expense_total = 0
        expenses_by_category = {}
        
        start = parse_date(start_date)
        end = parse_date(end_date)
        
        if not start or not end:
            return "❌ Ошибка в формате дат"
        
        sheet_name = None
        for name in ["Расходы", "расходы"]:
            if name in wb.sheetnames:
                sheet_name = name
                break
        
        if sheet_name:
            ws = wb[sheet_name]
            for row in range(2, ws.max_row + 1):
                date_cell = ws.cell(row=row, column=1).value
                category = ws.cell(row=row, column=2).value
                amount = ws.cell(row=row, column=4).value
                
                if date_cell and category and amount:
                    try:
                        row_date = parse_date(str(date_cell))
                        if row_date and start <= row_date <= end:
                            if isinstance(amount, str):
                                amount = amount.replace(' ', '').replace(',', '.').replace('₽', '').strip()
                            amount_val = float(amount)
                            expense_total += amount_val
                            expenses_by_category[category] = expenses_by_category.get(category, 0) + amount_val
                    except:
                        continue
        
        income_sheet = None
        for name in ["Доходы", "доходы"]:
            if name in wb.sheetnames:
                income_sheet = name
                break
        
        if income_sheet:
            ws = wb[income_sheet]
            for row in range(2, ws.max_row + 1):
                date_cell = ws.cell(row=row, column=1).value
                amount = ws.cell(row=row, column=3).value
                
                if date_cell and amount:
                    try:
                        row_date = parse_date(str(date_cell))
                        if row_date and start <= row_date <= end:
                            if isinstance(amount, str):
                                amount = amount.replace(' ', '').replace(',', '.').replace('₽', '').strip()
                            income_total += float(amount)
                    except:
                        continue
        
        balance = income_total - expense_total
        
        result = []
        result.append(f"💵 Доходы: {income_total:,.0f} ₽")
        result.append(f"💰 Расходы: {expense_total:,.0f} ₽")
        result.append(f"📊 Баланс: {balance:,.0f} ₽")
        
        if expenses_by_category:
            result.append("\n📊 По категориям:")
            sorted_cats = sorted(expenses_by_category.items(), key=lambda x: x[1], reverse=True)
            for cat, amt in sorted_cats[:5]:
                percent = (amt / expense_total * 100) if expense_total > 0 else 0
                result.append(f"  {cat}: {amt:,.0f} ₽ ({percent:.1f}%)")
        
        return "\n".join(result)
        
    except Exception as e:
        logger.error(f"Ошибка статистики за период: {e}")
        return f"❌ Ошибка: {str(e)[:100]}"

def compare_periods(start1: str, end1: str, start2: str, end2: str) -> str:
    """Сравнивает два периода"""
    if not all([start1, end1, start2, end2]):
        return "❌ Ошибка: не выбраны все даты"
    
    stats1 = get_statistics_period(start1, end1)
    stats2 = get_statistics_period(start2, end2)
    
    def extract_balance(text):
        import re
        match = re.search(r'Баланс: ([\d\s,]+) ₽', text)
        if match:
            return float(match.group(1).replace(' ', '').replace(',', ''))
        return 0
    
    def extract_income(text):
        match = re.search(r'Доходы: ([\d\s,]+) ₽', text)
        if match:
            return float(match.group(1).replace(' ', '').replace(',', ''))
        return 0
    
    def extract_expense(text):
        match = re.search(r'Расходы: ([\d\s,]+) ₽', text)
        if match:
            return float(match.group(1).replace(' ', '').replace(',', ''))
        return 0
    
    inc1 = extract_income(stats1)
    inc2 = extract_income(stats2)
    exp1 = extract_expense(stats1)
    exp2 = extract_expense(stats2)
    bal1 = extract_balance(stats1)
    bal2 = extract_balance(stats2)
    
    def format_change(val1, val2):
        if val1 == 0:
            return "∞"
        change = ((val2 - val1) / val1) * 100
        arrow = "📈" if change > 0 else "📉" if change < 0 else "➡️"
        return f"{change:+.1f}% {arrow}"
    
    result = []
    result.append(f"📈 Доходы:")
    result.append(f"  П1: {inc1:,.0f} ₽")
    result.append(f"  П2: {inc2:,.0f} ₽")
    result.append(f"  Изменение: {format_change(inc1, inc2)}")
    result.append("")
    result.append(f"💰 Расходы:")
    result.append(f"  П1: {exp1:,.0f} ₽")
    result.append(f"  П2: {exp2:,.0f} ₽")
    result.append(f"  Изменение: {format_change(exp1, exp2)}")
    result.append("")
    result.append(f"📊 Баланс:")
    result.append(f"  П1: {bal1:,.0f} ₽")
    result.append(f"  П2: {bal2:,.0f} ₽")
    result.append(f"  Изменение: {format_change(bal1, bal2)}")
    
    return "\n".join(result)

# ================== ЗАПУСК ТЕЛЕГРАМ БОТА ==================
async def setup_bot_commands(application: Application):
    """Установка команд для меню Telegram"""
    commands = [
        BotCommand("start", "🏠 Главное меню"),
        BotCommand("stats", "📊 Статистика"),
        BotCommand("help", "❓ Помощь"),
        BotCommand("ping", "🏓 Проверка"),
        BotCommand("debug", "🔧 Диагностика"),
    ]
    await application.bot.set_my_commands(commands)
    logger.info("✅ Команды меню установлены")

async def start_bot():
    """Запуск Telegram бота"""
    global bot_app
    
    logger.info("🤖 Инициализация Telegram бота...")
    
    for attempt in range(3):
        try:
            bot_app = Application.builder().token(BOT_TOKEN).build()
            logger.info(f"✅ Приложение создано (попытка {attempt + 1})")
            
            try:
                await bot_app.bot.delete_webhook(drop_pending_updates=True)
                logger.info("✅ Вебхук удален")
                await asyncio.sleep(2)
            except Exception as e:
                logger.warning(f"⚠️ Ошибка при очистке: {e}")
            
            await setup_bot_commands(bot_app)
            
            expense_conv = ConversationHandler(
                entry_points=[CallbackQueryHandler(button_callback, pattern="^method_")],
                states={
                    WAITING_EXPENSE_AMOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_expense_amount)]
                },
                fallbacks=[CommandHandler("cancel", cancel)],
                per_message=False
            )
            
            archive_expense_conv = ConversationHandler(
                entry_points=[CallbackQueryHandler(button_callback, pattern="^method_")],
                states={
                    WAITING_ARCHIVE_EXPENSE_AMOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_archive_expense_amount)]
                },
                fallbacks=[CommandHandler("cancel", cancel)],
                per_message=False
            )
            
            income_conv = ConversationHandler(
                entry_points=[CallbackQueryHandler(button_callback, pattern="^source_")],
                states={
                    WAITING_INCOME_AMOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_income_amount)]
                },
                fallbacks=[CommandHandler("cancel", cancel)],
                per_message=False
            )
            
            archive_income_conv = ConversationHandler(
                entry_points=[CallbackQueryHandler(button_callback, pattern="^source_")],
                states={
                    WAITING_ARCHIVE_INCOME_AMOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_archive_income_amount)]
                },
                fallbacks=[CommandHandler("cancel", cancel)],
                per_message=False
            )
            
            bot_app.add_handler(CommandHandler("start", start_command))
            bot_app.add_handler(CommandHandler("help", help_command))
            bot_app.add_handler(CommandHandler("stats", stats_command))
            bot_app.add_handler(CommandHandler("ping", ping_command))
            bot_app.add_handler(CommandHandler("debug", debug_command))
            bot_app.add_handler(CommandHandler("cancel", cancel))
            
            bot_app.add_handler(expense_conv)
            bot_app.add_handler(archive_expense_conv)
            bot_app.add_handler(income_conv)
            bot_app.add_handler(archive_income_conv)
            
            bot_app.add_handler(CallbackQueryHandler(button_callback))
            bot_app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_unknown))
            
            logger.info("✅ Все обработчики добавлены")
            
            await bot_app.initialize()
            await bot_app.start()
            
            await asyncio.sleep(3)
            
            await bot_app.updater.start_polling(
                poll_interval=0.5,
                timeout=30,
                drop_pending_updates=True,
                allowed_updates=['message', 'callback_query'],
                bootstrap_retries=5
            )
            
            logger.info("✅ Telegram бот успешно запущен!")
            return True
            
        except Exception as e:
            logger.error(f"💥 Ошибка при запуске бота (попытка {attempt + 1}): {e}")
            if attempt < 2:
                await asyncio.sleep(5)
            continue
    
    return False

# ================== АВТО-ПИНГ (ИСПРАВЛЕННАЯ ВЕРСИЯ) ==================
def start_auto_ping():
    """Запускает авто-пинг в отдельном потоке (как в проекте А)"""
    def ping_worker():
        time.sleep(30)
        
        # Собираем возможные URL для пинга
        possible_urls = []
        
        # Добавляем RENDER_URL из переменных окружения
        render_url = os.getenv("RENDER_URL", "").rstrip('/')
        if render_url:
            possible_urls.append(render_url)
        
        # Пробуем сгенерировать URL на основе hostname
        try:
            hostname = socket.gethostname()
            possible_urls.append(f"https://{hostname}.onrender.com")
        except:
            pass
        
        # Добавляем localhost как запасной вариант
        possible_urls.append(f"http://localhost:{PORT}")
        
        # Убираем пустые
        possible_urls = [url for url in possible_urls if url]
        
        logger.info(f"🧵 Авто-пинг запущен, пробуем URL: {possible_urls}")
        
        ping_count = 0
        while not shutdown_event.is_set():
            ping_count += 1
            
            for url in possible_urls:
                try:
                    response = requests.get(
                        f"{url}/health", 
                        timeout=15,
                        headers={"User-Agent": "Render-AutoPing/1.0"}
                    )
                    if response.status_code == 200:
                        logger.info(f"✅ Авто-пинг #{ping_count} успешен для {url}")
                        break
                except Exception as e:
                    logger.debug(f"Пинг #{ping_count} для {url} не удался: {e}")
                    continue
            else:
                logger.warning(f"⚠️ Авто-пинг #{ping_count}: все URL недоступны")
            
            # Пинг каждые 5 минут (300 секунд)
            for _ in range(300):
                if shutdown_event.is_set():
                    break
                time.sleep(1)
    
    thread = threading.Thread(target=ping_worker, daemon=True)
    thread.start()
    logger.info("✅ Поток авто-пинга создан")
    return thread

# ================== FASTAPI ЭНДПОИНТЫ ==================
app = FastAPI(title="Family Finance Bot")

@app.get("/")
@app.get("/health")
@app.get("/ping")
async def health_check():
    """Health check для Render"""
    return {
        "status": "healthy",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "bot_running": bool(bot_app and bot_app.running),
        "time_moscow": get_moscow_time(),
        "version": VERSION
    }

@app.get("/status")
async def status():
    """Статус системы"""
    return {
        "server": {
            "uptime": str(datetime.now(timezone.utc) - startup_time),
            "port": PORT,
            "startup_time": startup_time.isoformat()
        },
        "bot": {
            "initialized": bool(bot_app),
            "running": bool(bot_app and bot_app.running) if bot_app else False,
            "version": VERSION,
            "features": ["archive", "period_stats", "compare_periods"]
        }
    }

@app.on_event("startup")
async def startup_event():
    """Запуск при старте приложения"""
    logger.info("=" * 60)
    logger.info(f"🚀 ЗАПУСК ФИНАНСОВОГО БОТА v{VERSION}")
    logger.info("=" * 60)
    
    logger.info(f"✅ Токен бота: {'Найден' if BOT_TOKEN else 'ОТСУТСТВУЕТ!'}")
    logger.info(f"⏰ Время по Москве: {get_moscow_time()}")
    logger.info(f"📅 Дата: {get_current_date()}")
    logger.info(f"🌐 Порт: {PORT}")
    
    logger.info("=" * 60)
    
    global shutdown_event
    shutdown_event.clear()
    
    # Запускаем исправленный авто-пинг
    start_auto_ping()
    logger.info("🔧 Авто-пинг запущен")
    
    success = await start_bot()
    
    if success:
        logger.info("🎉 Все системы запущены!")
        logger.info("📅 Режим: Архивные платежи + Статистика + Сравнение")
    else:
        logger.error("💥 Не удалось запустить бота!")

@app.on_event("shutdown")
async def shutdown_event_handler():
    """Остановка при завершении"""
    logger.info("🛑 Завершение работы...")
    
    global shutdown_event
    shutdown_event.set()
    
    if bot_app:
        try:
            await bot_app.updater.stop()
            await bot_app.stop()
            await bot_app.shutdown()
            logger.info("✅ Telegram бот остановлен")
        except Exception as e:
            logger.error(f"❌ Ошибка при остановке бота: {e}")
    
    logger.info("👋 Сервер остановлен")

# ================== ТОЧКА ВХОДА ==================
def main():
    """Основная функция запуска"""
    logger.info(f"🌍 Запуск сервера на порту {PORT}...")
    
    uvicorn.run(
        "main:app",
        host="0.0.0.0",
        port=PORT,
        access_log=False,
        log_level="info",
        reload=False
    )

if __name__ == "__main__":
    main()
