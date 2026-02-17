"""
МОДУЛЬ РАБОТЫ С ЯНДЕКС.ДИСКОМ
Полная версия с функциями статистики
"""

import requests
from datetime import datetime
import logging
import time
from openpyxl import load_workbook
from config import YANDEX_TOKEN, PUBLIC_KEY, LOCAL_EXCEL_PATH

logger = logging.getLogger(__name__)


def download_from_yandex(max_retries=3):
    """Скачать файл с повторными попытками"""
    for attempt in range(max_retries):
        try:
            api_url = "https://cloud-api.yandex.net/v1/disk/public/resources/download"
            params = {"public_key": PUBLIC_KEY}
            response = requests.get(api_url, params=params, timeout=30)
            response.raise_for_status()
            
            download_url = response.json()["href"]
            response = requests.get(download_url, timeout=60)
            response.raise_for_status()
            
            with open(LOCAL_EXCEL_PATH, "wb") as f:
                f.write(response.content)
            
            logger.info("✅ Файл скачан с Яндекс.Диска")
            return True
            
        except Exception as e:
            logger.warning(f"Попытка {attempt + 1}/{max_retries} не удалась: {e}")
            time.sleep(2)
    
    logger.error("❌ Не удалось скачать файл после всех попыток")
    return False


def upload_to_yandex(max_retries=3):
    """Загрузить файл с повторными попытками"""
    for attempt in range(max_retries):
        try:
            headers = {"Authorization": f"OAuth {YANDEX_TOKEN}"}
            
            # Создаем папку Финансы (если её нет)
            folder_url = "https://cloud-api.yandex.net/v1/disk/resources"
            folder_params = {"path": "/Финансы"}
            requests.put(folder_url, headers=headers, params=folder_params)
            
            # Получаем ссылку для загрузки
            upload_url = "https://cloud-api.yandex.net/v1/disk/resources/upload"
            upload_params = {
                "path": "/Финансы/budget.xlsx",
                "overwrite": "true"
            }
            
            response = requests.get(upload_url, headers=headers, params=upload_params, timeout=30)
            response.raise_for_status()
            
            href = response.json()["href"]
            
            with open(LOCAL_EXCEL_PATH, "rb") as f:
                upload_response = requests.put(href, files={"file": f}, timeout=60)
                upload_response.raise_for_status()
            
            logger.info("✅ Файл загружен на Яндекс.Диск")
            return True
            
        except Exception as e:
            logger.warning(f"Попытка {attempt + 1}/{max_retries} не удалась: {e}")
            time.sleep(2)
    
    logger.error("❌ Не удалось загрузить файл после всех попыток")
    return False


def get_period():
    """Определить период по дню месяца"""
    day = datetime.now().day
    if day <= 9: 
        return "25-9"
    elif day <= 24: 
        return "10-24"
    else: 
        return "25-9"


def get_date():
    """Формат даты: ДД.ММ.ГГ"""
    return datetime.now().strftime("%d.%m.%y")


def clean_text(text):
    """Удалить эмодзи из текста"""
    if not text:
        return text
    parts = text.split(" ", 1)
    # Если первая часть - эмодзи, возвращаем вторую часть
    if len(parts) > 1 and parts[0].startswith(('🛒', '🏠', '🚗', '💳', '🌿', '💊', '🚬', '🐱', '🧹', '🎮', '🔨', '👕', '💇', '📦')):
        return parts[1]
    return text


def find_last_data_row(worksheet):
    """Находит последнюю строку с данными"""
    for row in range(worksheet.max_row, 1, -1):
        if worksheet.cell(row=row, column=1).value:
            return row
    return 1


def add_expense(category, amount, payer, payment_method):
    """Добавить расход"""
    try:
        if not download_from_yandex():
            return "❌ Не удалось скачать файл"
        
        wb = load_workbook(LOCAL_EXCEL_PATH)
        
        # Ищем лист с расходами
        sheet_name = None
        for name in ["Расходы", "расходы", "Лист1", "budget", "Sheet1"]:
            if name in wb.sheetnames:
                sheet_name = name
                break
        
        if not sheet_name:
            sheet_name = wb.sheetnames[0]  # первый доступный лист
            logger.info(f"Лист расходов не найден, используем: {sheet_name}")
        
        ws = wb[sheet_name]
        
        # Очищаем от эмодзи
        category_clean = clean_text(category)
        payer_clean = clean_text(payer)
        method_clean = clean_text(payment_method)
        
        # Находим последнюю строку с данными
        last_row = find_last_data_row(ws)
        new_row = last_row + 1
        
        # Добавляем данные
        ws.cell(row=new_row, column=1, value=get_date())        # A - Дата
        ws.cell(row=new_row, column=2, value=category_clean)    # B - Категория
        ws.cell(row=new_row, column=3, value="")                # C - Подкат
        ws.cell(row=new_row, column=4, value=float(amount))     # D - Сумма
        ws.cell(row=new_row, column=5, value=payer_clean)       # E - Кто
        ws.cell(row=new_row, column=6, value=get_period())      # F - Период
        ws.cell(row=new_row, column=7, value=method_clean)      # G - Способ
        
        # Сохраняем файл
        wb.save(LOCAL_EXCEL_PATH)
        
        if upload_to_yandex():
            return f"✅ Расход записан: {amount:,.0f} ₽, {category_clean}"
        else:
            return "⚠️ Расход записан локально, но не загружен в облако"
            
    except Exception as e:
        logger.error(f"Ошибка добавления расхода: {e}")
        return f"❌ Ошибка: {str(e)}"


def add_income(source, amount, payer):
    """Добавить доход"""
    try:
        if not download_from_yandex():
            return "❌ Не удалось скачать файл"
        
        wb = load_workbook(LOCAL_EXCEL_PATH)
        
        # Ищем лист с доходами
        sheet_name = None
        for name in ["Доходы", "доходы", "Лист1", "budget", "Sheet1"]:
            if name in wb.sheetnames:
                sheet_name = name
                break
        
        if not sheet_name:
            sheet_name = wb.sheetnames[0]  # первый доступный лист
            logger.info(f"Лист доходов не найден, используем: {sheet_name}")
        
        ws = wb[sheet_name]
        
        # Очищаем от эмодзи
        source_clean = clean_text(source)
        
        # Находим последнюю строку с данными
        last_row = find_last_data_row(ws)
        new_row = last_row + 1
        
        # Добавляем данные
        ws.cell(row=new_row, column=1, value=get_date())        # A - Дата
        ws.cell(row=new_row, column=2, value=source_clean)      # B - Источник
        ws.cell(row=new_row, column=3, value=float(amount))     # C - Сумма
        ws.cell(row=new_row, column=4, value=get_period())      # D - Период
        
        # Сохраняем файл
        wb.save(LOCAL_EXCEL_PATH)
        
        if upload_to_yandex():
            return f"✅ Доход записан: {amount:,.0f} ₽, {source_clean}"
        else:
            return "⚠️ Доход записан локально"
            
    except Exception as e:
        logger.error(f"Ошибка добавления дохода: {e}")
        return f"❌ Ошибка: {str(e)}"


def delete_last(sheet_name):
    """
    Удалить последнюю запись из указанного листа
    sheet_name: "Расходы" или "Доходы"
    """
    try:
        if not download_from_yandex():
            return "❌ Не удалось скачать файл"
        
        wb = load_workbook(LOCAL_EXCEL_PATH)
        
        # Находим лист
        target_sheet = None
        for name in [sheet_name, sheet_name.lower(), "Лист1", "budget", "Sheet1"]:
            if name in wb.sheetnames:
                target_sheet = name
                break
        
        if not target_sheet:
            return f"❌ Лист {sheet_name} не найден"
        
        ws = wb[target_sheet]
        
        # Находим последнюю строку с данными
        last_row = find_last_data_row(ws)
        
        if last_row <= 1:
            return "❌ Нет записей для удаления"
        
        # Сохраняем данные для сообщения
        date = ws.cell(row=last_row, column=1).value
        category = ws.cell(row=last_row, column=2).value
        amount = ws.cell(row=last_row, column=4).value
        
        # 🔧 Преобразование суммы в число для форматирования
        try:
            if amount is None:
                amount_float = 0
            else:
                # Если это строка, убираем пробелы и заменяем запятую на точку
                if isinstance(amount, str):
                    # Убираем все пробелы и заменяем запятую на точку
                    amount = amount.replace(' ', '').replace(',', '.')
                    # Если есть символ рубля, убираем его
                    amount = amount.replace('₽', '').replace('руб', '').strip()
                amount_float = float(amount)
                logger.info(f"Сумма для удаления преобразована: {amount} -> {amount_float}")
        except (ValueError, TypeError) as e:
            logger.warning(f"Не удалось преобразовать сумму '{amount}' в число: {e}")
            amount_float = 0
        
        # Удаляем строку
        ws.delete_rows(last_row)
        
        # Сохраняем файл
        wb.save(LOCAL_EXCEL_PATH)
        
        if upload_to_yandex():
            if sheet_name == "Расходы":
                return f"✅ Удалён расход: {date} | {category} | {amount_float:,.0f} ₽"
            else:
                return f"✅ Удалён доход: {date} | {category} | {amount_float:,.0f} ₽"
        else:
            return "⚠️ Запись удалена локально"
            
    except Exception as e:
        logger.error(f"Ошибка удаления: {e}")
        return f"❌ Ошибка удаления: {str(e)}"


# ========== НОВЫЕ ФУНКЦИИ СТАТИСТИКИ ==========

def get_statistics(by_categories=False, balance=False, period=None):
    """
    Получение статистики из Excel файла
    
    Параметры:
    - by_categories: True - расходы по категориям
    - balance: True - баланс (доходы - расходы)
    - period: "current" - текущий период (10-24)
             "previous" - предыдущий период (25-9)
             "all" - за всё время
    """
    try:
        if not download_from_yandex():
            return "❌ Не удалось скачать файл"
        
        wb = load_workbook(LOCAL_EXCEL_PATH, data_only=True)
        
        result = []
        
        # ===== СТАТИСТИКА ПО КАТЕГОРИЯМ РАСХОДОВ =====
        if by_categories:
            sheet_name = None
            for name in ["Расходы", "расходы", "Лист1", "budget"]:
                if name in wb.sheetnames:
                    sheet_name = name
                    break
            
            if sheet_name:
                ws = wb[sheet_name]
                
                # Собираем суммы по категориям
                categories = {}
                total = 0
                
                for row in range(2, ws.max_row + 1):
                    cat = ws.cell(row=row, column=2).value  # Категория
                    amount = ws.cell(row=row, column=4).value  # Сумма
                    
                    if cat and amount:
                        try:
                            # Преобразуем в число, если это строка
                            if isinstance(amount, str):
                                amount = amount.replace(' ', '').replace(',', '.').replace('₽', '').strip()
                            amount_val = float(amount)
                            categories[cat] = categories.get(cat, 0) + amount_val
                            total += amount_val
                        except (ValueError, TypeError):
                            continue
                
                # Сортируем по убыванию
                sorted_cats = sorted(categories.items(), key=lambda x: x[1], reverse=True)
                
                for cat, amt in sorted_cats[:10]:  # Топ-10
                    percent = (amt / total * 100) if total > 0 else 0
                    result.append(f"{cat}: {amt:,.0f} ₽ ({percent:.1f}%)")
                
                result.append(f"\n💰 Всего расходов: {total:,.0f} ₽")
            else:
                result.append("❌ Лист с расходами не найден")
        
        # ===== БАЛАНС (ДОХОДЫ - РАСХОДЫ) =====
        elif balance:
            income_total = 0
            expense_total = 0
            
            # Считаем доходы
            income_sheet = None
            for name in ["Доходы", "доходы"]:
                if name in wb.sheetnames:
                    income_sheet = name
                    break
            
            if income_sheet:
                ws = wb[income_sheet]
                for row in range(2, ws.max_row + 1):
                    amount = ws.cell(row=row, column=3).value  # Сумма в колонке C
                    if amount:
                        try:
                            if isinstance(amount, str):
                                amount = amount.replace(' ', '').replace(',', '.').replace('₽', '').strip()
                            income_total += float(amount)
                        except (ValueError, TypeError):
                            pass
            
            # Считаем расходы
            expense_sheet = None
            for name in ["Расходы", "расходы"]:
                if name in wb.sheetnames:
                    expense_sheet = name
                    break
            
            if expense_sheet:
                ws = wb[expense_sheet]
                for row in range(2, ws.max_row + 1):
                    amount = ws.cell(row=row, column=4).value  # Сумма в колонке D
                    if amount:
                        try:
                            if isinstance(amount, str):
                                amount = amount.replace(' ', '').replace(',', '.').replace('₽', '').strip()
                            expense_total += float(amount)
                        except (ValueError, TypeError):
                            pass
            
            balance_total = income_total - expense_total
            
            result.append(f"💵 Доходы: {income_total:,.0f} ₽")
            result.append(f"💰 Расходы: {expense_total:,.0f} ₽")
            result.append(f"📊 Баланс: {balance_total:,.0f} ₽")
        
        # ===== СТАТИСТИКА ЗА ПЕРИОД =====
        elif period:
            sheet_name = None
            for name in ["Расходы", "расходы"]:
                if name in wb.sheetnames:
                    sheet_name = name
                    break
            
            if sheet_name:
                ws = wb[sheet_name]
                
                period_map = {
                    "current": "10-24",
                    "previous": "25-9"
                }
                
                target_period = period_map.get(period) if period != "all" else None
                period_total = 0
                
                for row in range(2, ws.max_row + 1):
                    row_period = ws.cell(row=row, column=6).value  # Период
                    amount = ws.cell(row=row, column=4).value  # Сумма
                    
                    if period == "all" or (target_period and row_period == target_period):
                        if amount:
                            try:
                                if isinstance(amount, str):
                                    amount = amount.replace(' ', '').replace(',', '.').replace('₽', '').strip()
                                period_total += float(amount)
                            except (ValueError, TypeError):
                                pass
                
                if period == "all":
                    result.append(f"📅 Всего расходов за всё время: {period_total:,.0f} ₽")
                elif target_period == "10-24":
                    result.append(f"📅 Расходы за текущий период (10-24): {period_total:,.0f} ₽")
                elif target_period == "25-9":
                    result.append(f"📅 Расходы за предыдущий период (25-9): {period_total:,.0f} ₽")
            else:
                result.append("❌ Лист с расходами не найден")
        
        return "\n".join(result) if result else "❌ Нет данных для отображения"
        
    except Exception as e:
        logger.error(f"Ошибка получения статистики: {e}")
        return f"❌ Ошибка при подсчете статистики: {str(e)}"


def get_categories_summary():
    """Краткий обзор расходов по категориям (для быстрого отображения)"""
    return get_statistics(by_categories=True)


def get_balance_summary():
    """Краткий обзор баланса"""
    return get_statistics(balance=True)


def get_period_summary(period_type="current"):
    """Краткий обзор за период"""
    return get_statistics(period=period_type)
